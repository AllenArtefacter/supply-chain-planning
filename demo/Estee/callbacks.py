from dash import Dash, html, dcc, Input, Output, State, callback,Patch
import pandas as pd 
import numpy as np 
import re
import os 
from datetime import datetime as dt 
import dash_ag_grid as dag
import json
import plotly.graph_objects as go 
from datetime import timedelta
import numpy as np 
from itertools import permutations,product
from tqdm import tqdm 
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


from simulation import (
    generate_simulation,
    get_status,
    ttl_sales_through_rate,
    ttl_sales_shortage_rate,
    summarize_status,
    search_multi_process,
    simulate_one_trails
)
from elements import legend_description,search_result_aggrid,status_block

root_path = os.path.dirname(__file__)
history_path = os.path.join(root_path, 'data', 'sales_history.csv')
forecast_path = os.path.join(root_path, 'data', 'sales_forecast.csv')
hub_path = os.path.join(root_path, 'data', 'hub_stock.csv')

df_history = pd.read_csv(history_path)
df_forecast = pd.read_csv(forecast_path)
df_hub_stock = pd.read_csv(hub_path)

excel_path = os.path.join(root_path, 'data', 'allocation.xlsx')


@callback(
    Output("test1", "children",allow_duplicate=True),
    Output("status", "data",allow_duplicate=True),
    Output("allocation", "data",allow_duplicate=True),
    Input("search-results", "selectedRows"),
    State("config", "data"),
    prevent_initial_call=True,
    # prevent_initial_call='initial_duplicate'
)
def show_staus_from_config(row, config_data):
    # print("select", row)
    # print("conf1", config_data)
    # print("conf2", config_data2)
    if not row:
        return '', [],[]
        
    
    trial = row[0]['trial']
    # print(trial)
    
    config_ = config_data[trial]
    
    
    sku_ls = ['A', 'B', 'C']
    
    df_allocation = generate_simulation(
        config_, 
        sku_ls,'2023-01-14', '2023-01-15', df_history, df_forecast, df_hub_stock, [1,2,3,4]
    )
    #df_allocation['date'] = df_allocation['date'].apply(lambda x: x.strftime('%Y-%m-%d'))
    
    sales_through_rate = ttl_sales_through_rate(df_allocation)
    stockout_rate = ttl_sales_shortage_rate(df_allocation)
    df_status = summarize_status(df_allocation)
    
    df_status_ = df_status.copy()
    df_status_.columns = [i.strftime("%Y-%m-%d") if isinstance(i, dt) else i for i in df_status.columns]
    
    status_table = _status2table(df_status)
    
    status_table = status_block(status_table, config_, sales_through_rate, stockout_rate), 
    
    return (
        status_table, 
        df_status_.reset_index().to_dict('records'), 
        df_allocation.to_dict('records')
    )

@callback(
    Output("test1", "children",allow_duplicate=True),
    Output("status", "data",allow_duplicate=True),
    Output("allocation", "data",allow_duplicate=True),
    Input("history", "selectedRows"),
    State("config-records", "data"),
    prevent_initial_call=True,
    # prevent_initial_call='initial_duplicate'
)
def show_staus_from_history(row, config_data):
    # print("select", row)
    # print("conf1", config_data)
    # print("conf2", config_data2)
    
    if not row:
        return '', [],[]
    trial = row[0]['trial']
    print(trial)
    
    config_ = config_data[trial]
    
    
    sku_ls = ['A', 'B', 'C']
    
    df_allocation = generate_simulation(
        config_, 
        sku_ls,'2023-01-14', '2023-01-15', df_history, df_forecast, df_hub_stock, [1,2,3,4]
    )
    #df_allocation['date'] = df_allocation['date'].apply(lambda x: x.strftime('%Y-%m-%d'))
    
    sales_through_rate = ttl_sales_through_rate(df_allocation)
    stockout_rate = ttl_sales_shortage_rate(df_allocation)
    df_status = summarize_status(df_allocation)
    
    df_status_ = df_status.copy()
    df_status_.columns = [i.strftime("%Y-%m-%d") if isinstance(i, dt) else i for i in df_status.columns]
    
    status_table = _status2table(df_status)
    
    return (
        status_block(status_table, config_, sales_through_rate, stockout_rate), 
        df_status_.reset_index().to_dict('records'), 
        df_allocation.to_dict('records')
    )

@callback(
    Output("channel-parameter-display", "children"),
    Input("channel-parameter", "virtualRowData"),
    prevent_initial_call=True,
)
def display_channel_parameter_element(rows):
    dff = pd.DataFrame(rows)
    # return str(dff.columns)
    # print(dff.columns)
    order = dff['channel name'].tolist()
    
    return 'select order: '+' > '.join(order)


# @callback(
#     Output("test1", "children"),
#     Input("channel-parameter", "virtualRowData"),
#     # State("channel-parameter", "virtualRowData"),
# )
# def call_channel_parameter_element(rows):
#     dff = pd.DataFrame(rows)
    
#     return str(dff)

# @callback(
#     Output('optimize', 'n_clicks'),
#     Input('hold-buttom', 'modified_timestamp'),
#     State('hold-buttom', 'data')
# )
# def holding_button(_, data):
#     if data[0]:
#         return 0

@callback(
    [Output("config-search-records", "data"),
     Output("config", "data"),
     Output("search-results", "persistence")
     ],
    Input('optimize', 'n_clicks'),
    State("transportation1", "value"),
    State("transportation2", "value"),
    State("channel1-service-level", "value"),
    State("channel2-service-level", "value"),
    State("channel3-service-level", "value"),
    State("channel4-service-level", "value"),
    prevent_initial_call=True,
    blocking=True
)
def optimize_status_from_prameter(click, t1,  t2, c1,c2,c3,c4):
    print(click)
    if click >= 1:
        if len(t1):
            t1 = list(map(float, re.findall(r'\d+', str(t1))))
        else:
            return 'please select tranportation to hub', None
        if len(t2):
            t2 = list(map(float, re.findall(r'\d+', str(t2))))
        else:
            return 'please select tranportation to channel', None
            
        
        c1 = _generate_service_level(c1)
        c2 = _generate_service_level(c2)
        c3 = _generate_service_level(c3)
        c4 = _generate_service_level(c4)
        
        
        priority = [1,2,3,4]
        priority_permu = list(permutations(priority, 4))
        
    
        
        print(t1,t2,c1,c2,c3,c4,)
        combinations = list(
            product(   
                t1, 
                t2, 
                c1,
                c2,c3,c4,
                priority_permu
                )
        )
            
        config_spaces = [
            {   
                "trial":i,
                "lead_time_hud2channel": c[1],
                "lead_time_plant2hub": c[0],
                "safety_stock_factor" : {
                    "A" : 10,
                    "B" : 10, 
                    "C": 10
                },
                "channel_plan":{
                    "channel 1" :{
                        "priority" : c[-1][0],
                        "service_level" : c[2]
                    },        
                    "channel 2" :{
                        "priority" : c[-1][1],
                        "service_level" : c[3]
                    },
                    "channel 3" :{
                        "priority" : c[-1][2],
                        "service_level" : c[4]
                    },
                    "channel 4" :{
                        "priority" : c[-1][3],
                        "service_level" : c[5]
                    },
                }
            }
            for i, c in enumerate(combinations)
        ]
        
        search_rst = search_multi_process(
            config_spaces,
            simulate_one_trails,
            df_history=df_history, 
            df_forecast=df_forecast, 
            df_hub_stock=df_hub_stock
        )
        
        df_rst = pd.concat(search_rst, axis = 0)
        sales_through_threshold = df_rst['sales through'].quantile(.8)
        fill_rate_threshold = df_rst['fill rate'].quantile(.8)
        df_rst = df_rst[
            (df_rst['sales through'] >= sales_through_threshold ) | 
            (df_rst['fill rate'] >= fill_rate_threshold ) 
        ]
        
    
        return df_rst.to_dict('records'),config_spaces,True



def _generate_service_level(l):
    levels = []
    a = l[0]
    while a <= l[-1]:
        levels.append(a/100)
        a+= 10
    return levels
    
    
@callback(
    Output("history", "rowData"),
    Input("config-records-list", "modified_timestamp"),
    State("config-records-list", "data")
)
def update_history(_,data):
    # print('update',data)
    return data


@callback(
    Output("search-results", "rowData"),
    Input("config-search-records", "modified_timestamp"),
    State("config-search-records", "data")
)
def keep_search_result(_, data):
    return data
    
@callback(
    [Output("test1", "children"),
     Output("status", "data"),
     Output("allocation", "data"),
     Output("config-records", "data"),
     Output("config-records-list", "data"),
     ],
    Input('submit1', 'n_clicks'),
    State("transportation1", "value"),
    State("transportation2", "value"),
    State("channel-parameter", "virtualRowData"),
    State("config-records", "data"),
    State("config-records-list", "data"),
    prevent_initial_call=True,
    # prevent_initial_callbacks='initial_duplicate'
)
def get_status_from_prameter(click, t1,  t2, channel_rows, r,rl):
    
    if click > 0:
        if not t1:
            return "please select tranportation to hub"
        if not t2:
            return "please select tranportation to channel"
        

        print('click', click)
        t1 = float(re.findall(r'\d+', str(t1))[0])
        t2 = float(re.findall(r'\d+', str(t2))[0])
        
        df_channels = pd.DataFrame(channel_rows)
        
        df_channels['service_level'] = df_channels['service level']/100
        df_channels['priority'] = range(len(df_channels))

        channel_dict = df_channels.set_index('channel name')[['priority', 'service_level']].T.to_dict()
        
        config = {
            "lead_time_hud2channel": t2,
            "lead_time_plant2hub": t1,
            "safety_stock_factor" : {
                "A" : 10,
                "B" : 10, 
                "C": 10
            },
            "channel_plan":channel_dict
        }
        # return str(config)
        sku_ls = ['A', 'B', 'C']
        
        df_allocation = generate_simulation(
            config, 
            sku_ls,'2023-01-14', '2023-01-15', df_history, df_forecast, df_hub_stock, [1,2,3,4]
        )
        #df_allocation['date'] = df_allocation['date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        
        sales_through_rate = ttl_sales_through_rate(df_allocation)
        stockout_rate = ttl_sales_shortage_rate(df_allocation)
        df_status = summarize_status(df_allocation)
        
        df_status_ = df_status.copy()
        df_status_.columns = [i.strftime("%Y-%m-%d") if isinstance(i, dt) else i for i in df_status.columns]
        
        status_table = _status2table(df_status)
        
        status_table = status_block(status_table, config, sales_through_rate, stockout_rate), 
        
        r.append(config)
        channel_plan = config['channel_plan']
        channel_order = sorted(channel_plan, key=lambda x: channel_plan[x]['priority'])

        order = ''
        for o in channel_order:
            order += f" {o}({int(config['channel_plan'][o]['service_level'] * 100)}) "
            
        rl.append({
                "trial": len(r)-1,
                "sales through": [sales_through_rate],
                "fill rate": [1- stockout_rate],
                "transportation days (to hub)" : [config['lead_time_plant2hub']],
                "transportation days (to channel)" : [config['lead_time_hud2channel']],
                "safety_stock_factor": [config['safety_stock_factor']],
                "service level and order": [order]
            })
        
        df_rl = pd.DataFrame(rl)
        return (
            status_table, 
            df_status_.reset_index().to_dict('records'), 
            df_allocation.to_dict('records'),
            r, 
            rl,
        )

@callback(
    Output("status-details", "children"),
    Input("status-table", "cellClicked"),
    State("status", "data"),
    State("allocation", "data"),
    prevent_initial_call=True
)
def display_status_details(cell, status_records, allocation_records):
    if cell:
        df_status = pd.DataFrame(status_records)
        
        
        df_allocation = pd.DataFrame(allocation_records)
        r_idx = cell['rowIndex']
        c_idx = cell['colId']
        c_idx = dt.strptime(c_idx, '%Y %b %d').strftime('%Y-%m-%d') if len(c_idx) ==11 and len(c_idx.split(' '))==3 else None
        
        
        sku, channel  = df_status[['sku_name', 'channel']].values[r_idx]
        
        fig = _plot_details(df_allocation, sku, channel,c_idx)

        return html.Div(
            dcc.Graph(figure=fig,style = {'height':'190px', 'margin':'5px'}),
            style = {'heigh':'200px','margin-top':'1px', 'border':'solid 0.5px grey','border-radius':'2px'}
        )

def _status2table(df_status):
    df_status_ = df_status.reset_index()
    # df_status_['sales_through'] = df_status_['sales_through'].apply('')
    df_status_ = df_status_.rename(columns = {'sku_name':'sku', 'sales_through':'sales through', 'stockout': 'stockout rate'})
    
    datetime_cols = [i.strftime("%Y %b %d") for i in df_status_.columns[4:]]
    df_status_.columns = list(df_status_.columns[:4]) + datetime_cols

    # columnDef = [
    #     {
    #         'headerName': 'Info',
    #         'children':[]
    #     },
    #     {
    #         'headerName': 'Date',
    #         'children':[]
    #     }
    # ]
    
    comment_columnDef = [
        {'field': 'sku','cellStyle': {'font-size': '12rem'}, 'maxWidth':50,},
        {'field': 'channel','cellStyle': {'font-size': '12rem'}, 'maxWidth':100},
        {'field': 'sales through','cellStyle': {'font-size': '12rem'}, 'maxWidth':100,'valueFormatter': {"function":"params.value == null ? '' :  d3.format(',.1%')(params.value)" }},
        {'field': 'stockout rate','cellStyle': {'font-size': '12rem'}, 'maxWidth':100,'valueFormatter': {"function":"params.value == null ? '' :  d3.format(',.1%')(params.value)" }},
    ]
    
    date_columnDef =  [
        {'field': c, 'cellStyle': {'font-size': '12rem'},'maxWidth':40,'sortable':False,
            'cellStyle': {
            "styleConditions": [
                {
                    "condition": "params.value == 'full'",
                    "style": {"backgroundColor": "#98D8AA", "color":"#98D8AA"},
                },
                {
                    "condition": "['hub', 'manufacture'].includes(params.value)",
                    "style": {"backgroundColor": "#F7D060", "color":"#F7D060"},
                },
                {
                    "condition": "params.value == 'risk'",
                    "style": {"backgroundColor": "#E99497", "color":"#E99497"},
                },
                {
                    "condition": "params.value == 'near stockout'",
                    "style": {"backgroundColor": "#FF6363", "color":"#FF6363"},
                },
                {
                    "condition": "params.value == 'stockout'",
                    "style": {"backgroundColor": "#577B8D", "color":"#577B8D"},
                },
            ]
            
            }
            }
        for c in df_status_.columns[4:]
    ]
    columnDef = comment_columnDef + date_columnDef
    status_table = dag.AgGrid(
            # className= 'ag-status-table',
            id='status-table',
            rowData=df_status_.to_dict('records'),
            columnDefs=columnDef,
            defaultColDef={"filter": False},
            columnSize="sizeToFit",
            dashGridOptions= {"rowHeight": 30},
            columnSizeOptions={
                'defaultMinWidth': 80,
                'columnLimits': [{'key': c, 'minWidth': 50} for c in datetime_cols],
            },
            style= {
                'height':400, 'font-size':'12rem', 'font-family':'Century Gothic', 'margin':'0px'
            }
    )
    
    return status_table


def _plot_details(df_allocation,sku, channel, date_id=None):
    fig = go.Figure()
    df_allocation_selected = df_allocation.query(f"sku_name == '{sku}' & channel == '{channel}'")
    df_allocation_selected
    fig.add_hline(y = df_allocation_selected['safety_stock'].values[0], line = dict(color = '#E99497', width = 1, dash = 'dot'))
    fig.add_trace(
        go.Waterfall(
        name = "20", orientation = "v",
        x = df_allocation_selected['date'],
        textposition = "outside",
        # text = df_allocation_selected['status'],
        # decreasing = {"marker":{"color":"#F7D060"}},
        increasing = {"marker":{"color":"#F7D060"}},
        y = [df_allocation_selected['stock'].values[0]] + df_allocation_selected['stock'].diff().tolist()[1:],
        connector = {"line":{"color":"rgb(63, 63, 63)"}},
    ))

    fig.add_annotation(
        x = 1, y = df_allocation_selected['safety_stock'].values[0],
        xref='paper', yref='y',
        xanchor='left',
        align='left',
        text= 'Safety<br>Stock<br>Level',
        showarrow=False,
        font = {'color': '#E99497'}
    )

    max_stock = df_allocation_selected['stock'].max()
    for i, d in df_allocation_selected.reset_index().iterrows():
        if d['status'] not in ['risk', 'full', 'near stockout'] or i == 0:
            fig.add_annotation(
                x = d['date'], y = d['stock'] + max_stock*0.02,
                # xanchor='',
                align='center',
                text= 'current<br>stock'if i==0 else d['status'].replace('stockout', 'stock<br>out'),
                font = dict(color = 'red' if d['status'] == 'stockout' else 'black'),
                showarrow=True,arrowhead = 2
            )

    fig.update_layout(
        template = 'simple_white',showlegend = False, margin = {'t':30, 'l':0,'b':0, 'r':50},
        font = dict(size = 11, family = 'Century Gothic'), title = f'{sku} {channel}'
    )
    fig.update_xaxes(mirror = True,)
    fig.update_yaxes(mirror = True, title = 'stock', range = [0, max_stock * 1.5])

    if date_id:
        x0 = dt.strptime(date_id, '%Y-%m-%d') - timedelta(days = 0.5)
        x1 = dt.strptime(date_id, '%Y-%m-%d') + timedelta(days = 0.5)
        fig.add_shape(type="rect",
            x0=x0, y0=0, x1=x1, y1=max_stock * 1.5,
            line=dict(
                color="RoyalBlue",
                width=2,
            ),
            fillcolor="LightSkyBlue",
        )

    return fig


@callback(
    Output("history", "dashGridOptions"),
    Input("history", "selectedRows"),
)
def row_pinning_top(selected_rows):
    grid_option_patch = Patch()
    grid_option_patch["pinnedTopRowData"] = selected_rows
    return grid_option_patch


@callback(
    Output("search-results", "dashGridOptions"),
    Input("search-results", "selectedRows"),
)
def row_pinning_top2(selected_rows):
    grid_option_patch = Patch()
    grid_option_patch["pinnedTopRowData"] = selected_rows
    return grid_option_patch

@callback(
    Output('download',"data"),
    Input('download-button', 'n_clicks'),
    State('allocation', 'data'),
    State('status','data'),
    State('status-config','data'),
    prevent_initial_call=True
)
def download_excel(clicks,allocation, status, config):
    if clicks > 0:
        # xw.App().visible = False
        df_status = pd.DataFrame(status).reset_index(drop=True)
        df_allocation = pd.DataFrame(allocation)
        df_config = pd.DataFrame(config)
        # wb = xw.Book(excel_path)  # Connects to the active instance of Excel
        # sheet = wb.sheets['Sheet1']
        # sheet.range("A:L")[1:,:].clear_contents()
        # sheet['A1'].options(index=False).value = df_allocation
        # sheet = wb.sheets['Sheet2']
        # sheet.range('A3').value = df_status
        # wb.save('allocation_download.xlsx')
        # wb.close()

        wb = openpyxl.load_workbook(excel_path)
        sheet = wb['allocation']
        for a in sheet['A1':'M1000']: #you can set the range here 
            for cell in a:
                cell.value = None
        data = dataframe_to_rows(df_allocation, index=False, header=True)
            
        for r, row in enumerate(data, start=1):
            for c, value in enumerate(row, start=1):
                sheet.cell(row=r, column=c).value = value
                
                
        sheet = wb['status']
        data = dataframe_to_rows(df_status, index=False, header=True)

        for r, row in enumerate(data, start=1):
            for c, value in enumerate(row, start=1):
                sheet.cell(row=r+2, column=c).value = value
        sheet.cell(column =1, row = 2).value = ttl_sales_through_rate(df_allocation)
        sheet.cell(column =3, row = 2).value = ttl_sales_shortage_rate(df_allocation)
        
        sheet = wb['config']
        data = dataframe_to_rows(df_config, index=False, header=True)

        for r, row in enumerate(data, start=1):
            for c, value in enumerate(row, start=1):
                sheet.cell(row=r, column=c).value = str(value)
        
        wb.save('allocation_download.xlsx')
        
        return dcc.send_file(
            "allocation_download.xlsx"
        )
    
    
@callback(
    Output("test4", "children"),
    Input("allocation", "data"),
    Input("status", "data"),
    prevent_initial_call=True
)
def alert_summary(allocation, statue):
    
    df_allocation = pd.DataFrame(allocation)
    
    df_summary = (
        df_allocation
        .rename(columns= {'sku_name':'sku'})
        .pivot_table(
            index= ['sku', 'channel' ],
            columns= 'status',
            values = 'date',
            aggfunc= 'count'
        )
        .reset_index()
    )
    
    df_start_date = (
        df_allocation
            .rename(columns= {'sku_name':'sku'})
            .groupby(['sku', 'channel'], as_index=False)
            .agg({'date':'min'})
            .rename(columns= {'date':'est start date'})
    )
    df_start_date['est start date'] =  df_start_date['est start date'].apply(
        lambda x: pd.to_datetime(x).strftime("%b %d %Y")
    )

    df_end_date = (
        df_allocation
        .rename(columns= {'sku_name':'sku'})
        .groupby(['sku', 'channel'], as_index=False)
        .agg({'date':'max'})
        .rename(columns= {'date':'est end date'})
    )
    df_end_date['est end date'] =  df_end_date['est end date'].apply(
        lambda x: pd.to_datetime(x).strftime("%b %d %Y")
    )
    
    
    df_summary = (
        df_summary
        .merge(df_start_date, on = ['sku', 'channel'])
        .merge(df_end_date, on = ['sku', 'channel'])
    )
    
    print(df_summary.columns)
    
    columnDef = [
        {'field': 'sku', 'cellStyle': {'font-size': '12rem'}, 'maxWidth':100},
        {'field': 'channel', 'cellStyle': {'font-size': '12rem'},'maxWidth':100},
        {'field': 'full', 'cellStyle': {'font-size': '12rem'},'maxWidth':100,},
        {'field': 'risk', 'cellStyle': {'font-size': '12rem'},'maxWidth':100,},
        {'field': 'near stockout', 'cellStyle': {'font-size': '12rem'},'maxWidth':100,},
        {'field': 'stockout', 'cellStyle': {'font-size': '12rem'},'maxWidth':100,},
        {'field': 'est start date', 'cellStyle': {'font-size': '12rem'},'maxWidth':150},
        {'field': 'est end date', 'cellStyle': {'font-size': '12rem'},'maxWidth':150},
    ]
    
    summary_table = dag.AgGrid(
            # className= 'ag-status-table',
            # id='status-table',
            rowData=df_summary.to_dict('records'),
            columnDefs=columnDef,
            defaultColDef={"filter": False},
            columnSize="sizeToFit",
            dashGridOptions= {"rowHeight": 30},
            columnSizeOptions={
                'defaultMinWidth': 80,
                # 'columnLimits': [{'key': c, 'minWidth': 50} for c in datetime_cols],
            },
            style= {
                'height':400, 'font-size':'12rem', 'font-family':'Century Gothic', 'margin':'0px'
            }
    )
    return [summary_table]